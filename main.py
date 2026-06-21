from readiness import build_readiness_report
from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from bodycomp import extract_inbody_data, validate_inbody_data, build_trend_summary
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, JSONResponse
import httpx
import os
import json
from datetime import datetime, timedelta
from dotenv import load_dotenv
import anthropic
import asyncpg
import asyncio
from contextlib import asynccontextmanager

load_dotenv()

# ─── DB ────────────────────────────────────────────────────────────────────────
async def get_db():
    return await asyncpg.connect(os.getenv("DATABASE_URL"))

async def init_db():
    conn = await get_db()
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS tokens (
            provider TEXT PRIMARY KEY,
            access_token TEXT,
            refresh_token TEXT,
            expires_at BIGINT
        );
        CREATE TABLE IF NOT EXISTS activities (
            id TEXT PRIMARY KEY,
            provider TEXT,
            date DATE,
            type TEXT,
            data JSONB,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS daily_metrics (
            date DATE PRIMARY KEY,
            whoop_recovery FLOAT,
            whoop_hrv FLOAT,
            whoop_rhr FLOAT,
            whoop_sleep_score FLOAT,
            whoop_strain FLOAT,
            garmin_body_battery FLOAT,
            garmin_stress FLOAT,
            garmin_steps INT,
            garmin_hrv FLOAT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS body_composition (
            id SERIAL PRIMARY KEY,
            exam_date DATE NOT NULL,
            exam_time TEXT,
            weight_kg FLOAT,
            body_water_l FLOAT,
            protein_kg FLOAT,
            minerals_kg FLOAT,
            fat_mass_kg FLOAT,
            skeletal_muscle_mass_kg FLOAT,
            bmi FLOAT,
            body_fat_pct FLOAT,
            inbody_score INT,
            fat_free_mass_kg FLOAT,
            basal_metabolic_rate_kcal INT,
            visceral_fat_level INT,
            waist_hip_ratio FLOAT,
            smi FLOAT,
            created_at TIMESTAMP DEFAULT NOW()
        );
    """)
    await conn.close()

@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    import asyncio
    asyncio.create_task(daily_scheduler())
    yield

app = FastAPI(title="Personal Coach Agent", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
analyze_jobs: dict = {}

# ─── WHOOP OAuth ───────────────────────────────────────────────────────────────
WHOOP_CLIENT_ID     = os.getenv("WHOOP_CLIENT_ID")
WHOOP_CLIENT_SECRET = os.getenv("WHOOP_CLIENT_SECRET")
WHOOP_REDIRECT_URI  = os.getenv("BASE_URL") + "/auth/whoop/callback"
WHOOP_AUTH_URL      = "https://api.prod.whoop.com/oauth/oauth2/auth"
WHOOP_TOKEN_URL     = "https://api.prod.whoop.com/oauth/oauth2/token"
WHOOP_API_BASE      = "https://api.prod.whoop.com/developer/v2"

# ─── Garmin Health API ─────────────────────────────────────────────────────────
GARMIN_CLIENT_ID     = os.getenv("GARMIN_CLIENT_ID")
GARMIN_CLIENT_SECRET = os.getenv("GARMIN_CLIENT_SECRET")
GARMIN_REDIRECT_URI  = os.getenv("BASE_URL") + "/auth/garmin/callback"
GARMIN_AUTH_URL      = "https://connect.garmin.com/oauthConfirm"
GARMIN_TOKEN_URL     = "https://connect.garmin.com/oauth-service/oauth/access_token"
GARMIN_API_BASE      = "https://apis.garmin.com/wellness-api/rest"

# ─── Strava OAuth ──────────────────────────────────────────────────────────────
STRAVA_CLIENT_ID     = os.getenv("STRAVA_CLIENT_ID")
STRAVA_CLIENT_SECRET = os.getenv("STRAVA_CLIENT_SECRET")
STRAVA_REDIRECT_URI  = os.getenv("BASE_URL") + "/auth/strava/callback"
STRAVA_AUTH_URL      = "https://www.strava.com/oauth/authorize"
STRAVA_TOKEN_URL     = "https://www.strava.com/oauth/token"
STRAVA_API_BASE      = "https://www.strava.com/api/v3"

# ─── COROS API ─────────────────────────────────────────────────────────────────
COROS_CLIENT_ID     = os.getenv("COROS_CLIENT_ID")
COROS_CLIENT_SECRET = os.getenv("COROS_CLIENT_SECRET")
COROS_REDIRECT_URI  = os.getenv("BASE_URL") + "/auth/coros/callback"
COROS_AUTH_URL      = "https://open.coros.com/oauth2/authorize"
COROS_TOKEN_URL     = "https://open.coros.com/oauth2/accesstoken"
COROS_API_BASE      = "https://open.coros.com"

# ─── Token helpers ─────────────────────────────────────────────────────────────
async def save_token(provider: str, access: str, refresh: str, expires_in: int):
    conn = await get_db()
    expires_at = int(datetime.now().timestamp()) + expires_in
    await conn.execute("""
        INSERT INTO tokens (provider, access_token, refresh_token, expires_at)
        VALUES ($1,$2,$3,$4)
        ON CONFLICT (provider) DO UPDATE
        SET access_token=$2, refresh_token=$3, expires_at=$4
    """, provider, access, refresh, expires_at)
    await conn.close()

async def get_token(provider: str) -> dict | None:
    conn = await get_db()
    row = await conn.fetchrow("SELECT * FROM tokens WHERE provider=$1", provider)
    await conn.close()
    if not row:
        return None
    return dict(row)

async def refresh_whoop_token(refresh_token: str) -> dict:
    async with httpx.AsyncClient() as client:
        r = await client.post(WHOOP_TOKEN_URL, data={
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "client_id": WHOOP_CLIENT_ID,
            "client_secret": WHOOP_CLIENT_SECRET,
        })
        return r.json()

async def refresh_strava_token(refresh_token: str) -> dict:
    async with httpx.AsyncClient() as client:
        r = await client.post(STRAVA_TOKEN_URL, data={
            "client_id": STRAVA_CLIENT_ID,
            "client_secret": STRAVA_CLIENT_SECRET,
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        })
        return r.json()

async def valid_token(provider: str) -> str | None:
    t = await get_token(provider)
    if not t:
        return None
    now = int(datetime.now().timestamp())
    if t["expires_at"] - now < 300:  # refresh se expira em <5min
        if provider == "whoop":
            new = await refresh_whoop_token(t["refresh_token"])
            await save_token(provider, new["access_token"], new["refresh_token"], new["expires_in"])
            return new["access_token"]
        elif provider == "strava":
            new = await refresh_strava_token(t["refresh_token"])
            await save_token(provider, new["access_token"], new["refresh_token"], new["expires_in"])
            return new["access_token"]
    return t["access_token"]

# ─── Auth routes ───────────────────────────────────────────────────────────────
@app.get("/auth/whoop")
async def auth_whoop():
    import secrets
    state = secrets.token_hex(8)
    url = (f"{WHOOP_AUTH_URL}?client_id={WHOOP_CLIENT_ID}"
           f"&redirect_uri={WHOOP_REDIRECT_URI}&response_type=code"
           f"&scope=read:recovery read:sleep read:workout read:body_measurement offline"
           f"&state={state}")
    return RedirectResponse(url)

@app.get("/auth/whoop/callback")
async def auth_whoop_callback(code: str):
    async with httpx.AsyncClient() as client:
        r = await client.post(WHOOP_TOKEN_URL, data={
            "code": code,
            "grant_type": "authorization_code",
            "redirect_uri": WHOOP_REDIRECT_URI,
            "client_id": WHOOP_CLIENT_ID,
            "client_secret": WHOOP_CLIENT_SECRET,
        })
    data = r.json()
    await save_token("whoop", data["access_token"], data["refresh_token"], data["expires_in"])
    return {"status": "WHOOP conectado!", "provider": "whoop"}

@app.get("/auth/garmin")
async def auth_garmin():
    url = (f"{GARMIN_AUTH_URL}?client_id={GARMIN_CLIENT_ID}"
           f"&redirect_uri={GARMIN_REDIRECT_URI}&response_type=code&scope=")
    return RedirectResponse(url)

@app.get("/auth/garmin/callback")
async def auth_garmin_callback(code: str):
    async with httpx.AsyncClient() as client:
        r = await client.post(GARMIN_TOKEN_URL, data={
            "code": code,
            "grant_type": "authorization_code",
            "redirect_uri": GARMIN_REDIRECT_URI,
            "client_id": GARMIN_CLIENT_ID,
            "client_secret": GARMIN_CLIENT_SECRET,
        })
    data = r.json()
    await save_token("garmin", data["access_token"], data.get("refresh_token",""), data.get("expires_in", 3600))
    return {"status": "Garmin conectado!", "provider": "garmin"}

@app.get("/auth/strava")
async def auth_strava():
    url = (f"{STRAVA_AUTH_URL}?client_id={STRAVA_CLIENT_ID}"
           f"&redirect_uri={STRAVA_REDIRECT_URI}&response_type=code"
           f"&approval_prompt=auto&scope=read,activity:read_all")
    return RedirectResponse(url)

@app.get("/auth/strava/callback")
async def auth_strava_callback(code: str):
    async with httpx.AsyncClient() as client:
        r = await client.post(STRAVA_TOKEN_URL, data={
            "code": code,
            "client_id": STRAVA_CLIENT_ID,
            "client_secret": STRAVA_CLIENT_SECRET,
            "grant_type": "authorization_code",
        })
    data = r.json()
    await save_token("strava", data["access_token"], data["refresh_token"], data["expires_in"])
    return {"status": "Strava conectado!", "provider": "strava"}

@app.get("/auth/coros")
async def auth_coros():
    url = (f"{COROS_AUTH_URL}?client_id={COROS_CLIENT_ID}"
           f"&redirect_uri={COROS_REDIRECT_URI}&response_type=code&state=coros")
    return RedirectResponse(url)

@app.get("/auth/coros/callback")
async def auth_coros_callback(code: str):
    async with httpx.AsyncClient() as client:
        r = await client.post(COROS_TOKEN_URL, data={
            "code": code,
            "client_id": COROS_CLIENT_ID,
            "client_secret": COROS_CLIENT_SECRET,
            "redirect_uri": COROS_REDIRECT_URI,
            "grant_type": "authorization_code",
        })
    data = r.json()
    token_data = data.get("data", data)
    await save_token("coros", token_data["access_token"], token_data.get("refresh_token",""), token_data.get("expires_in", 7200))
    return {"status": "COROS conectado!", "provider": "coros"}

# ─── Status das conexões ───────────────────────────────────────────────────────
@app.get("/status")
async def status():
    result = {}
    for provider in ["whoop", "garmin", "strava", "coros"]:
        t = await get_token(provider)
        result[provider] = {
            "connected": t is not None,
            "expires_at": datetime.fromtimestamp(t["expires_at"]).isoformat() if t else None
        }
    return result

# ─── Coleta de dados WHOOP ─────────────────────────────────────────────────────
async def fetch_whoop_data(days: int = 7) -> dict:
    token = await valid_token("whoop")
    if not token:
        return {}
    headers = {"Authorization": f"Bearer {token}"}
    start = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%dT00:00:00.000Z")

    async with httpx.AsyncClient() as client:
        r_rec = await client.get(f"{WHOOP_API_BASE}/recovery", headers=headers,
                                  params={"start": start, "limit": days})
        r_sleep = await client.get(f"{WHOOP_API_BASE}/activity/sleep", headers=headers,
                                    params={"start": start, "limit": days})
        r_work = await client.get(f"{WHOOP_API_BASE}/activity/workout", headers=headers,
                                   params={"start": start, "limit": days * 2})

    return {
        "recovery": r_rec.json().get("records", []) if r_rec.status_code == 200 else [],
        "sleep": r_sleep.json().get("records", []) if r_sleep.status_code == 200 else [],
        "workouts": r_work.json().get("records", []) if r_work.status_code == 200 else [],
        "_debug": {
            "recovery_status": r_rec.status_code,
            "sleep_status": r_sleep.status_code,
            "workout_status": r_work.status_code,
        }
    }

# ─── Coleta de dados Garmin ────────────────────────────────────────────────────
async def fetch_garmin_data(days: int = 7) -> dict:
    token = await valid_token("garmin")
    if not token:
        return {}
    headers = {"Authorization": f"Bearer {token}"}
    upload_start = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    upload_end = datetime.now().strftime("%Y-%m-%d")

    async with httpx.AsyncClient() as client:
        r_daily = await client.get(
            f"{GARMIN_API_BASE}/dailies",
            headers=headers,
            params={"uploadStartTimeInSeconds": upload_start, "uploadEndTimeInSeconds": upload_end}
        )
        r_sleep = await client.get(
            f"{GARMIN_API_BASE}/sleeps",
            headers=headers,
            params={"uploadStartTimeInSeconds": upload_start, "uploadEndTimeInSeconds": upload_end}
        )

    return {
        "dailies": r_daily.json() if r_daily.status_code == 200 else [],
        "sleep": r_sleep.json() if r_sleep.status_code == 200 else [],
    }

# ─── Coleta de dados Strava ────────────────────────────────────────────────────
async def fetch_strava_data(days: int = 14) -> list:
    token = await valid_token("strava")
    if not token:
        return []
    headers = {"Authorization": f"Bearer {token}"}
    after = int((datetime.now() - timedelta(days=days)).timestamp())

    async with httpx.AsyncClient() as client:
        r = await client.get(f"{STRAVA_API_BASE}/athlete/activities",
                             headers=headers,
                             params={"after": after, "per_page": 30})
    return r.json() if r.status_code == 200 else []

# ─── Coleta de dados COROS ────────────────────────────────────────────────────
async def fetch_coros_data(days: int = 14) -> list:
    token = await valid_token("coros")
    if not token:
        return []
    headers = {"Authorization": f"Bearer {token}"}
    start = int((datetime.now() - timedelta(days=days)).timestamp())
    end   = int(datetime.now().timestamp())

    async with httpx.AsyncClient() as client:
        r = await client.get(f"{COROS_API_BASE}/v2/coros/sport/list",
                             headers=headers,
                             params={"startTime": start, "endTime": end, "size": 50})
    data = r.json()
    return data.get("data", {}).get("dataList", []) if r.status_code == 200 else []

# ─── Sync endpoint ─────────────────────────────────────────────────────────────
@app.post("/sync")
async def sync_all(days: int = 7):
    whoop   = await fetch_whoop_data(days)
    garmin  = await fetch_garmin_data(days)
    strava  = await fetch_strava_data(days)
    coros   = await fetch_coros_data(days)
    return {
        "whoop": whoop,
        "garmin": garmin,
        "strava": strava,
        "coros": coros,
        "synced_at": datetime.now().isoformat()
    }

# ─── Análise com Claude ────────────────────────────────────────────────────────
@app.post("/analyze")
async def analyze(days: int = 7):
    import uuid
    job_id = str(uuid.uuid4())
    analyze_jobs[job_id] = {"status": "processando", "result": None}
    asyncio.create_task(run_analysis_job(job_id, days))
    return {"job_id": job_id, "status": "processando"}


async def run_analysis_job(job_id: str, days: int):
    try:
        data = await sync_all(days)
        report = build_readiness_report(data.get("whoop", {}))
        conn = await get_db()
        bodycomp_rows = await conn.fetch("""
            SELECT * FROM body_composition ORDER BY exam_date DESC LIMIT 10
        """)
        await conn.close()
        bodycomp_history = [dict(r) for r in bodycomp_rows]
        for h in bodycomp_history:
            h["exam_date"] = h["exam_date"].isoformat() if h["exam_date"] else None
            h["created_at"] = h["created_at"].isoformat() if h["created_at"] else None
        bodycomp_trend = build_trend_summary(bodycomp_history)
        ai = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        prompt = f"""Você é um treinador pessoal especialista em corrida e performance atlética.
Analise os dados abaixo dos últimos {days} dias e retorne um JSON com esta estrutura exata:

{{
  "resumo": "análise geral em 2-3 frases",
  "estado_recovery": "verde|amarelo|vermelho",
  "carga_semana": "baixa|moderada|alta|muito_alta",
  "insights": ["insight 1", "insight 2", "insight 3"],
  "alerta": "alerta importante se houver, ou null",
  "composicao_corporal": "comentário sobre a tendência de peso/gordura/músculo se houver dados, ou null",
  "plano_semana": [
    {{
      "dia": "Segunda",
      "data": "YYYY-MM-DD",
      "tipo": "Corrida fácil|Intervalado|Tempo run|Long run|Força|Descanso ativo|Descanso",
      "descricao": "detalhes do treino",
      "duracao_min": 45,
      "distancia_km": 8.0,
      "zona_fc": "Z1-Z2",
      "intensidade": "leve|moderada|alta",
      "justificativa": "por que esse treino hoje"
    }}
  ]
}}

Composição corporal (histórico de exames InBody, mais recente primeiro):
{json.dumps(bodycomp_trend, ensure_ascii=False, default=str)[:1500]}

Dados disponíveis:
WHOOP: {json.dumps(data.get('whoop', {}), ensure_ascii=False, default=str)[:3000]}
WHOOP: {json.dumps(data.get('whoop', {}), ensure_ascii=False, default=str)[:3000]}
GARMIN: {json.dumps(data.get('garmin', {}), ensure_ascii=False, default=str)[:2000]}
STRAVA (atividades recentes): {json.dumps(data.get('strava', [])[:5], ensure_ascii=False, default=str)[:2000]}
COROS (corridas recentes): {json.dumps(data.get('coros', [])[:5], ensure_ascii=False, default=str)[:2000]}

Retorne APENAS o JSON, sem markdown, sem explicações."""

        msg = ai.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )

        raw = msg.content[0].text.strip()
        try:
            analysis = json.loads(raw)
        except Exception:
            analysis = {"raw": raw}

        analyze_jobs[job_id] = {
            "status": "concluido",
            "result": {
                "analysis": analysis,
                "readiness_report": report,
                "raw_data": data,
                "generated_at": datetime.now().isoformat(),
            }
        }
    except Exception as e:
        analyze_jobs[job_id] = {"status": "erro", "error": str(e)}


@app.get("/analyze/status/{job_id}")
async def analyze_status(job_id: str):
    job = analyze_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job não encontrado")
    return job

# ─── Gerar planilha Excel ──────────────────────────────────────────────────────
@app.get("/export/excel")
async def export_excel(days: int = 7):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from fastapi.responses import StreamingResponse

    result = await analyze(days)
    analysis = result["analysis"]
    plan = analysis.get("plano_semana", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plano de Treino"

    # Cores
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    verde_fill  = PatternFill("solid", fgColor="2D6A4F")
    amarelo_fill= PatternFill("solid", fgColor="E9C46A")
    vermelho_fill=PatternFill("solid", fgColor="C1121F")
    alt_fill    = PatternFill("solid", fgColor="F8F9FA")

    tipo_cores = {
        "Descanso": "ADB5BD",
        "Descanso ativo": "CED4DA",
        "Corrida fácil": "A8DADC",
        "Long run": "457B9D",
        "Intervalado": "E63946",
        "Tempo run": "F4A261",
        "Força": "8338EC",
    }

    bold_white = Font(bold=True, color="FFFFFF", size=11)
    bold_dark  = Font(bold=True, color="1A1A2E", size=11)
    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(cell, value, fill=header_fill, font=bold_white, align="center"):
        cell.value = value
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = border

    def dcell(cell, value, fill=None, align="left", bold=False):
        cell.value = value
        if fill:
            cell.fill = fill
        cell.font = Font(bold=bold, color="1A1A2E", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = border

    # ── Título ──
    ws.merge_cells("A1:I1")
    hcell(ws["A1"], "🏃 PLANO DE TREINO PERSONALIZADO", font=Font(bold=True, color="FFFFFF", size=14))
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} • Análise dos últimos {days} dias"
    ws["A2"].font = Font(color="6C757D", italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Resumo ──
    ws.merge_cells("A4:I4")
    hcell(ws["A4"], "ANÁLISE GERAL", font=Font(bold=True, color="FFFFFF", size=11))

    ws.merge_cells("A5:I5")
    ws["A5"].value = analysis.get("resumo", "")
    ws["A5"].font = Font(size=10, color="1A1A2E")
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A5"].fill = PatternFill("solid", fgColor="E8F4F8")
    ws.row_dimensions[5].height = 45

    # Status badges
    estado = analysis.get("estado_recovery", "verde")
    carga  = analysis.get("carga_semana", "moderada")
    estado_fill = {"verde": verde_fill, "amarelo": amarelo_fill, "vermelho": vermelho_fill}.get(estado, verde_fill)

    ws.merge_cells("A7:D7")
    hcell(ws["A7"], f"RECOVERY: {estado.upper()}", fill=estado_fill)

    ws.merge_cells("E7:I7")
    hcell(ws["E7"], f"CARGA DA SEMANA: {carga.upper()}", fill=PatternFill("solid", fgColor="264653"))
    ws.row_dimensions[7].height = 25

    # Insights
    ws.merge_cells("A9:I9")
    hcell(ws["A9"], "INSIGHTS")
    for i, insight in enumerate(analysis.get("insights", []), start=10):
        ws.merge_cells(f"A{i}:I{i}")
        ws[f"A{i}"].value = f"• {insight}"
        ws[f"A{i}"].font = Font(size=10)
        ws[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"A{i}"].fill = PatternFill("solid", fgColor="F0F7FF") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[i].height = 30

    alerta = analysis.get("alerta")
    row_after = 10 + len(analysis.get("insights", []))
    if alerta:
        ws.merge_cells(f"A{row_after}:I{row_after}")
        hcell(ws[f"A{row_after}"], f"⚠️ {alerta}", fill=vermelho_fill)
        ws.row_dimensions[row_after].height = 30
        row_after += 1

    # ── Plano semanal ──
    row = row_after + 1
    ws.merge_cells(f"A{row}:I{row}")
    hcell(ws[f"A{row}"], "PLANO DA SEMANA")
    ws.row_dimensions[row].height = 28
    row += 1

    headers = ["Dia", "Data", "Tipo de Treino", "Descrição", "Duração", "Distância", "Zona FC", "Intensidade", "Justificativa"]
    cols    = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    for col, h in zip(cols, headers):
        hcell(ws[f"{col}{row}"], h)
    ws.row_dimensions[row].height = 22
    row += 1

    for i, treino in enumerate(plan):
        tipo = treino.get("tipo", "")
        cor_hex = tipo_cores.get(tipo, "FFFFFF")
        tipo_fill = PatternFill("solid", fgColor=cor_hex)
        bg_fill = PatternFill("solid", fgColor="F8F9FA") if i % 2 == 0 else None

        dcell(ws[f"A{row}"], treino.get("dia", ""), fill=bg_fill, bold=True)
        dcell(ws[f"B{row}"], treino.get("data", ""), fill=bg_fill, align="center")
        hcell(ws[f"C{row}"], tipo, fill=tipo_fill, font=Font(bold=True, color="1A1A2E" if cor_hex in ["A8DADC","E9C46A","ADB5BD","CED4DA"] else "FFFFFF"))
        dcell(ws[f"D{row}"], treino.get("descricao", ""), fill=bg_fill)
        dcell(ws[f"E{row}"], f"{treino.get('duracao_min', '')} min", fill=bg_fill, align="center")
        dcell(ws[f"F{row}"], f"{treino.get('distancia_km', '')} km" if treino.get("distancia_km") else "–", fill=bg_fill, align="center")
        dcell(ws[f"G{row}"], treino.get("zona_fc", ""), fill=bg_fill, align="center")
        dcell(ws[f"H{row}"], treino.get("intensidade", ""), fill=bg_fill, align="center")
        dcell(ws[f"I{row}"], treino.get("justificativa", ""), fill=bg_fill)
        ws.row_dimensions[row].height = 50
        row += 1

    # Larguras das colunas
    widths = [12, 12, 18, 40, 10, 12, 10, 12, 40]
    for col, w in zip(cols, widths):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"treino_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

# ─── Health check ──────────────────────────────────────────────────────────────
@app.get("/")
async def root():
    return {"status": "ok", "service": "Personal Coach Agent"}

# ─── Relatório diário de sono e recovery ──────────────────────────────────────
async def generate_daily_report() -> dict:
    days = 2
    whoop  = await fetch_whoop_data(days)
    garmin = await fetch_garmin_data(days)
    strava = await fetch_strava_data(days)
    coros  = await fetch_coros_data(days)

    ai = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

    prompt = f"""Você é um treinador pessoal e especialista em recuperação atlética.
Analise os dados abaixo e gere um relatório diário completo em JSON com esta estrutura exata:

{{
  "data": "DD/MM/YYYY",
  "saudacao": "Bom dia, Bruno!",
  "resumo_executivo": "2-3 frases sobre o estado geral de hoje",
  "sono": {{
    "duracao_horas": 7.5,
    "qualidade": "boa|regular|ruim",
    "nota": 85,
    "estagio_rem_min": 90,
    "estagio_deep_min": 60,
    "estagio_light_min": 120,
    "disturbios": 3,
    "analise": "análise detalhada do sono desta noite",
    "comparativo": "comparado com sua média recente"
  }},
  "recovery": {{
    "score": 78,
    "status": "verde|amarelo|vermelho",
    "hrv": 65.0,
    "hrv_tendencia": "subindo|estavel|caindo",
    "fc_repouso": 52,
    "analise": "análise do recovery de hoje"
  }},
  "strain": {{
    "score_ontem": 12.5,
    "nivel": "leve|moderado|alto|muito alto",
    "analise": "como a carga de ontem afeta hoje"
  }},
  "recomendacao_treino": {{
    "tipo": "Corrida fácil|Intervalado|Tempo run|Long run|Força|Descanso ativo|Descanso",
    "descricao": "descrição detalhada do treino recomendado",
    "duracao_min": 45,
    "distancia_km": 8.0,
    "zona_fc": "Z1-Z2",
    "justificativa": "por que esse treino baseado nos dados de hoje"
  }},
  "dicas_do_dia": ["dica 1", "dica 2", "dica 3"],
  "alerta": "alerta importante se houver, ou null"
}}

Dados disponíveis:
WHOOP (recovery, sono, workouts): {json.dumps(whoop, ensure_ascii=False, default=str)[:3000]}
GARMIN (dailies, sono): {json.dumps(garmin, ensure_ascii=False, default=str)[:2000]}
STRAVA: {json.dumps(strava[:3], ensure_ascii=False, default=str)[:1000]}
COROS: {json.dumps(coros[:3], ensure_ascii=False, default=str)[:1000]}

Retorne APENAS o JSON, sem markdown, sem explicações."""

    msg = ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=3000,
        messages=[{"role": "user", "content": prompt}]
    )

    raw = msg.content[0].text.strip()
    try:
        return json.loads(raw)
    except:
        return {"raw": raw}


def build_email_html(report: dict) -> str:
    sono    = report.get("sono", {})
    rec     = report.get("recovery", {})
    strain  = report.get("strain", {})
    treino  = report.get("recomendacao_treino", {})
    dicas   = report.get("dicas_do_dia", [])
    alerta  = report.get("alerta")

    status_cores = {"verde": "#00C9A7", "amarelo": "#F4C542", "vermelho": "#E63946"}
    rec_cor = status_cores.get(rec.get("status", "verde"), "#00C9A7")

    qualidade_cores = {"boa": "#00C9A7", "regular": "#F4C542", "ruim": "#E63946"}
    sono_cor = qualidade_cores.get(sono.get("qualidade", "boa"), "#00C9A7")

    alerta_html = f"""
    <div style="background:#2D1B1B;border-left:4px solid #E63946;padding:14px 18px;border-radius:8px;margin-bottom:20px;">
      <p style="color:#FF6B6B;margin:0;font-size:14px;">⚠️ {alerta}</p>
    </div>""" if alerta else ""

    dicas_html = "".join([f'<li style="color:#C8D0E0;font-size:14px;margin-bottom:8px;">{d}</li>' for d in dicas])

    return f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0D0D1A;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
  <div style="max-width:600px;margin:0 auto;padding:32px 20px;">

    <!-- Header -->
    <div style="background:linear-gradient(135deg,#1A1A2E,#16213E);border-radius:16px;padding:28px;margin-bottom:20px;border:1px solid #2A2A4A;">
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
        <span style="font-size:32px;">🏃</span>
        <div>
          <h1 style="color:#F0F0F5;margin:0;font-size:22px;font-weight:700;">Coach Agent</h1>
          <p style="color:#6C757D;margin:0;font-size:13px;">Relatório diário • {report.get("data", "")}</p>
        </div>
      </div>
      <p style="color:#C8D0E0;margin:0;font-size:15px;line-height:1.6;">{report.get("resumo_executivo", "")}</p>
    </div>
    
    {alerta_html}

    <!-- Sono -->
    <div style="background:#1A1A2E;border-radius:12px;padding:22px;margin-bottom:16px;border:1px solid #2A2A4A;">
      <h2 style="color:#F0F0F5;margin:0 0 16px;font-size:16px;font-weight:600;">😴 Sono desta noite</h2>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:16px;">
        <div style="background:#0D0D1A;border-radius:8px;padding:12px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 4px;font-size:11px;text-transform:uppercase;">Duração</p>
          <p style="color:#F0F0F5;margin:0;font-size:20px;font-weight:700;">{sono.get("duracao_horas", "–")}h</p>
        </div>
        <div style="background:#0D0D1A;border-radius:8px;padding:12px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 4px;font-size:11px;text-transform:uppercase;">Qualidade</p>
          <p style="color:{sono_cor};margin:0;font-size:20px;font-weight:700;">{sono.get("qualidade","–").upper()}</p>
        </div>
        <div style="background:#0D0D1A;border-radius:8px;padding:12px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 4px;font-size:11px;text-transform:uppercase;">Nota</p>
          <p style="color:#4361EE;margin:0;font-size:20px;font-weight:700;">{sono.get("nota","–")}</p>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:14px;">
        <div style="background:#0D0D1A;border-radius:6px;padding:8px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 2px;font-size:10px;">REM</p>
          <p style="color:#8338EC;margin:0;font-size:14px;font-weight:600;">{sono.get("estagio_rem_min","–")} min</p>
        </div>
        <div style="background:#0D0D1A;border-radius:6px;padding:8px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 2px;font-size:10px;">Profundo</p>
          <p style="color:#4361EE;margin:0;font-size:14px;font-weight:600;">{sono.get("estagio_deep_min","–")} min</p>
        </div>
        <div style="background:#0D0D1A;border-radius:6px;padding:8px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 2px;font-size:10px;">Leve</p>
          <p style="color:#00C9A7;margin:0;font-size:14px;font-weight:600;">{sono.get("estagio_light_min","–")} min</p>
        </div>
      </div>
      <p style="color:#8892A4;margin:0;font-size:13px;line-height:1.6;">{sono.get("analise","")}</p>
      <p style="color:#6C757D;margin:8px 0 0;font-size:12px;font-style:italic;">{sono.get("comparativo","")}</p>
    </div>

    <!-- Recovery -->
    <div style="background:#1A1A2E;border-radius:12px;padding:22px;margin-bottom:16px;border:1px solid {rec_cor}33;">
      <h2 style="color:#F0F0F5;margin:0 0 16px;font-size:16px;font-weight:600;">⚡ Recovery</h2>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:14px;">
        <div style="background:#0D0D1A;border-radius:8px;padding:12px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 4px;font-size:11px;text-transform:uppercase;">Score</p>
          <p style="color:{rec_cor};margin:0;font-size:24px;font-weight:700;">{rec.get("score","–")}%</p>
        </div>
        <div style="background:#0D0D1A;border-radius:8px;padding:12px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 4px;font-size:11px;text-transform:uppercase;">HRV</p>
          <p style="color:#F0F0F5;margin:0;font-size:24px;font-weight:700;">{rec.get("hrv","–")}</p>
          <p style="color:#6C757D;margin:0;font-size:10px;">{rec.get("hrv_tendencia","")}</p>
        </div>
        <div style="background:#0D0D1A;border-radius:8px;padding:12px;text-align:center;">
          <p style="color:#6C757D;margin:0 0 4px;font-size:11px;text-transform:uppercase;">FC repouso</p>
          <p style="color:#F0F0F5;margin:0;font-size:24px;font-weight:700;">{rec.get("fc_repouso","–")}</p>
        </div>
      </div>
      <p style="color:#8892A4;margin:0;font-size:13px;line-height:1.6;">{rec.get("analise","")}</p>
    </div>

    <!-- Treino recomendado -->
    <div style="background:#1A1A2E;border-radius:12px;padding:22px;margin-bottom:16px;border:1px solid #4361EE33;">
      <h2 style="color:#F0F0F5;margin:0 0 16px;font-size:16px;font-weight:600;">🎯 Treino de hoje</h2>
      <div style="background:#4361EE22;border-radius:8px;padding:14px;margin-bottom:12px;">
        <p style="color:#4361EE;margin:0 0 6px;font-size:16px;font-weight:700;">{treino.get("tipo","")}</p>
        <p style="color:#C8D0E0;margin:0;font-size:13px;">{treino.get("descricao","")}</p>
      </div>
      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px;">
        <div style="text-align:center;">
          <p style="color:#6C757D;margin:0 0 2px;font-size:10px;text-transform:uppercase;">Duração</p>
          <p style="color:#F0F0F5;margin:0;font-size:14px;font-weight:600;">{treino.get("duracao_min","–")} min</p>
        </div>
        <div style="text-align:center;">
          <p style="color:#6C757D;margin:0 0 2px;font-size:10px;text-transform:uppercase;">Distância</p>
          <p style="color:#F0F0F5;margin:0;font-size:14px;font-weight:600;">{treino.get("distancia_km","–")} km</p>
        </div>
        <div style="text-align:center;">
          <p style="color:#6C757D;margin:0 0 2px;font-size:10px;text-transform:uppercase;">Zona FC</p>
          <p style="color:#F0F0F5;margin:0;font-size:14px;font-weight:600;">{treino.get("zona_fc","–")}</p>
        </div>
      </div>
      <p style="color:#6C757D;margin:0;font-size:12px;font-style:italic;">💡 {treino.get("justificativa","")}</p>
    </div>

    <!-- Dicas -->
    <div style="background:#1A1A2E;border-radius:12px;padding:22px;margin-bottom:20px;border:1px solid #2A2A4A;">
      <h2 style="color:#F0F0F5;margin:0 0 12px;font-size:16px;font-weight:600;">💡 Dicas do dia</h2>
      <ul style="margin:0;padding-left:20px;">
        {dicas_html}
      </ul>
    </div>

    <!-- Footer -->
    <p style="color:#3A3A5A;text-align:center;font-size:11px;margin:0;">Coach Agent • WHOOP + Garmin + COROS + Strava → Claude AI</p>
  </div>
</body>
</html>"""


async def send_daily_email(report: dict):
    import httpx
    api_key = os.getenv("RESEND_API_KEY")
    if not api_key:
        return

    html = build_email_html(report)
    data_str = report.get("data", datetime.now().strftime("%d/%m/%Y"))

    async with httpx.AsyncClient() as client:
        await client.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "from": "Coach Agent <onboarding@resend.dev>",
                "to": ["dobruno@gmail.com"],
                "subject": f"🏃 Relatório diário — {data_str}",
                "html": html,
            }
        )


# ─── Scheduler (10h Brasília = 13h UTC) ───────────────────────────────────────
async def daily_scheduler():
    while True:
        now = datetime.utcnow()
        target = now.replace(hour=12, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        wait_seconds = (target - now).total_seconds()
        await asyncio.sleep(wait_seconds)
        try:
            report = await generate_daily_report()
            await send_daily_email(report)
        except Exception as e:
            print(f"Erro no relatório diário: {e}")
            

# ─── Endpoints manuais de relatório ───────────────────────────────────────────
@app.post("/report/daily")
async def report_daily():
    report = await generate_daily_report()
    await send_daily_email(report)
    return {"status": "email enviado", "report": report}

@app.get("/report/preview")
async def report_preview():
    report = await generate_daily_report()
    return report

@app.get("/dashboard")
async def dashboard():
    from fastapi.responses import FileResponse
    return FileResponse("dashboard.html")

@app.get("/debug/claude")
async def debug_claude():
    import time
    t0 = time.time()
    ai = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    msg = ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=50,
        messages=[{"role": "user", "content": "responda apenas: ok"}]
    )
    t1 = time.time()
    return {"seconds": round(t1 - t0, 2), "response": msg.content[0].text}

@app.get("/debug/analyze_steps")
async def debug_analyze_steps(days: int = 7):
    import time
    steps = {}

    t0 = time.time()
    whoop = await fetch_whoop_data(days)
    steps["whoop"] = round(time.time() - t0, 2)

    t0 = time.time()
    garmin = await fetch_garmin_data(days)
    steps["garmin"] = round(time.time() - t0, 2)

    t0 = time.time()
    strava = await fetch_strava_data(days)
    steps["strava"] = round(time.time() - t0, 2)

    t0 = time.time()
    coros = await fetch_coros_data(days)
    steps["coros"] = round(time.time() - t0, 2)

    t0 = time.time()
    report = build_readiness_report(whoop)
    steps["readiness"] = round(time.time() - t0, 2)

    return {"steps": steps}

@app.get("/debug/analyze_full")
async def debug_analyze_full(days: int = 7):
    import time
    t0 = time.time()
    data = await sync_all(days)
    t1 = time.time()

    report = build_readiness_report(data.get("whoop", {}))
    t2 = time.time()

    ai = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

    prompt = f"""Você é um treinador pessoal especialista em corrida e performance atlética.
Analise os dados abaixo dos últimos {days} dias e retorne um JSON com esta estrutura exata:

{{
  "resumo": "análise geral em 2-3 frases",
  "estado_recovery": "verde|amarelo|vermelho",
  "carga_semana": "baixa|moderada|alta|muito_alta",
  "insights": ["insight 1", "insight 2", "insight 3"],
  "alerta": "alerta importante se houver, ou null",
  "plano_semana": [
    {{
      "dia": "Segunda",
      "data": "YYYY-MM-DD",
      "tipo": "Corrida fácil|Intervalado|Tempo run|Long run|Força|Descanso ativo|Descanso",
      "descricao": "detalhes do treino",
      "duracao_min": 45,
      "distancia_km": 8.0,
      "zona_fc": "Z1-Z2",
      "intensidade": "leve|moderada|alta",
      "justificativa": "por que esse treino hoje"
    }}
  ]
}}

Dados disponíveis:
WHOOP: {json.dumps(data.get('whoop', {}), ensure_ascii=False, default=str)[:3000]}
GARMIN: {json.dumps(data.get('garmin', {}), ensure_ascii=False, default=str)[:2000]}
STRAVA (atividades recentes): {json.dumps(data.get('strava', [])[:5], ensure_ascii=False, default=str)[:2000]}
COROS (corridas recentes): {json.dumps(data.get('coros', [])[:5], ensure_ascii=False, default=str)[:2000]}

Retorne APENAS o JSON, sem markdown, sem explicações."""

    t3 = time.time()
    msg = ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}]
    )
    t4 = time.time()

    return {
        "sync_seconds": round(t1 - t0, 2),
        "readiness_seconds": round(t2 - t1, 2),
        "prompt_build_seconds": round(t3 - t2, 2),
        "prompt_length_chars": len(prompt),
        "claude_call_seconds": round(t4 - t3, 2),
        "response_preview": msg.content[0].text[:200],
    }

@app.post("/upload/inbody")
async def upload_inbody(file: UploadFile = File(...)):
    image_bytes = await file.read()
    media_type = file.content_type or "image/jpeg"

    extracted = extract_inbody_data(image_bytes, media_type)
    if "error" in extracted:
        raise HTTPException(status_code=422, detail=extracted)

    is_valid, warnings = validate_inbody_data(extracted)

    exam_date_str = extracted.get("exam_date")
    exam_date_obj = None
    if exam_date_str:
        try:
            exam_date_obj = datetime.strptime(exam_date_str, "%Y-%m-%d").date()
        except ValueError:
            warnings.append(f"exam_date inválido: {exam_date_str}")

    conn = await get_db()
    await conn.execute("""
        INSERT INTO body_composition (
            exam_date, exam_time, weight_kg, body_water_l, protein_kg,
            minerals_kg, fat_mass_kg, skeletal_muscle_mass_kg, bmi,
            body_fat_pct, inbody_score, fat_free_mass_kg,
            basal_metabolic_rate_kcal, visceral_fat_level,
            waist_hip_ratio, smi
        ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)
    """,
        exam_date_obj, extracted.get("exam_time"),
        extracted.get("weight_kg"), extracted.get("body_water_l"),
        extracted.get("protein_kg"), extracted.get("minerals_kg"),
        extracted.get("fat_mass_kg"), extracted.get("skeletal_muscle_mass_kg"),
        extracted.get("bmi"), extracted.get("body_fat_pct"),
        extracted.get("inbody_score"), extracted.get("fat_free_mass_kg"),
        extracted.get("basal_metabolic_rate_kcal"), extracted.get("visceral_fat_level"),
        extracted.get("waist_hip_ratio"), extracted.get("smi")
    )
    await conn.close()

    return {
        "status": "salvo" if is_valid else "salvo_com_alertas",
        "extracted": extracted,
        "warnings": warnings
    }


@app.get("/bodycomp/history")
async def bodycomp_history(limit: int = 20):
    conn = await get_db()
    rows = await conn.fetch("""
        SELECT * FROM body_composition
        ORDER BY exam_date DESC
        LIMIT $1
    """, limit)
    await conn.close()

    history = [dict(r) for r in rows]
    for h in history:
        h["exam_date"] = h["exam_date"].isoformat() if h["exam_date"] else None
        h["created_at"] = h["created_at"].isoformat() if h["created_at"] else None

    trend = build_trend_summary(history)
    return {"history": history, "trend": trend}
